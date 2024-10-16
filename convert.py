import os
import re
import json
from openpyxl import load_workbook
import app.constants as constants

if __name__ == "__main__":
    path = os.path.join(constants.BUILD_SOURCES_DIR, constants.mitigations_mapping_dir)
    file = f"{path}/mappings.xlsx"

    workbook = load_workbook(file, data_only=True)
    
    main_sheet = workbook["Blue Team Guide"]
    uses_sheet = workbook["MITRE Uses"]
    mitre_sheet = workbook["MITRE Controls"]
    nist_sheet = workbook["NIST Controls"]
    ism_sheet = workbook["ISM Controls"]

    mitigations = {}
    uses = {}
    isms = {}
    nists = {}

    # mitre control uses
    overarching_mitigation = ""
    overarching_technique = ""

    for row_i, row in enumerate(uses_sheet.iter_rows()):
        if row_i == 1: continue

        technique = ""

        # 0: id, 1: sub-technique, 3: use
        for i, col in enumerate(row):
            col = str(col.value)

            if i == 0:
                mitigation = re.findall(r".*?\((M[0-9]{4})\)", col)
                technique = re.findall(r"(T[0-9]{4})", col)

                if len(technique) != 0: overarching_technique = technique[0]
                if len(mitigation) != 0: overarching_mitigation = mitigation[0]
            if i == 1 and row_i != 0:
                technique = overarching_technique

                if col != "None":
                    technique = f"{overarching_technique}.{col.split('.')[1]}"
            if i == 3 and row_i != 0 and col != "None" and col != "Use":
                if not overarching_mitigation in uses:
                    uses[overarching_mitigation] = {}
                
                uses[overarching_mitigation][technique] = { "use": col }

    # main sheet with all information
    overarching_technique = ""

    for row_i, row in enumerate(main_sheet.iter_rows()):
        if row_i == 0: continue

        full_technique = ""

        for i, col in enumerate(row):
            col = str(col.value)

            if i == 0:
                technique = re.findall(r".*\((T[0-9]{4})\)", col)
                subtechnique = re.findall(r".*\((\.[0-9]{3})\)", col)

                if len(technique) != 0:
                    full_technique = technique[0]
                    overarching_technique = technique[0]

                if len(subtechnique) != 0:
                    full_technique = overarching_technique + subtechnique[0]
            if i == 6:
                ism_controls = re.findall(r"ISM-[0-9]{4}", col)

                for ism in ism_controls:
                    if ism not in isms:
                        isms[ism] = {}
                    
                    isms[ism][full_technique] = { 'use': None }
            if i == 7:
                nist_controls = re.findall(r".*?([A-Z]{2}-[0-9]{1,2})", col)

                for nist in nist_controls:
                    if nist not in nists:
                        nists[nist] = {}

                    nists[nist][full_technique] = { 'use': None }

    # mitre controls
    for i, rows in enumerate(mitre_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "MITRE", "techniques": [] }

        # 0: id, 1: name, 2: description
        for i, row in enumerate(rows):
            if i == 1: values["name"] = row.value
            if i == 2: values["description"] = row.value
            if i == 0: 
                mitigations[row.value] = values
                
                if row.value in uses:
                    mitigations[row.value]["techniques"] = uses[row.value]

    # nist controls
    for i, rows in enumerate(nist_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "NIST", "techniques": [] }

        # 0: id, 1: name, 2: description
        for i, row in enumerate(rows):

            # not a nist sub-technique
            if re.match(r"[A-Z]{2}-[0-9]{1,2}\([0-9]{1,2}\)", str(row.value)): continue
            
            if i == 1: values["name"] = row.value
            if i == 2: values["description"] = row.value
            if i == 0: 
                mitigations[row.value] = values

                if row.value in nists:
                    mitigations[row.value]["techniques"] = nists[row.value]

    # ism controls
    for i, rows in enumerate(ism_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "ISM", "techniques": [] }

        # 0: id, 1: name, 2: description
        for i, row in enumerate(rows):
            if i == 1: values["name"] = row.value
            if i == 2: values["description"] = row.value
            if i == 0:
                mitigations[row.value] = values

                if row.value in isms:
                    mitigations[row.value]["techniques"] = isms[row.value]

    with open(f"{path}/mappings-v15.1.json", "w") as outfile: 
        json.dump(mitigations, outfile)